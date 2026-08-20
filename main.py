import io
import requests
from fastapi import FastAPI, Form, HTTPException, Header, Security
from fastapi.security import APIKeyHeader
from fastapi.responses import StreamingResponse
import pdfplumber
import openpyxl

# ReportLab bileşenleri (Excel'den kaliteli PDF üretmek için)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

app = FastAPI(title="Akilli Doc Converter API")

# --- API KEY ALTYAPISI ---
API_KEY_NAME = "X-API-Key"
api_key_header = APIKeyHeader(name=API_KEY_NAME, auto_error=False)

def verify_api_key(x_api_key: str = Header(None)):
    # Şimdilik API Key opsiyoneldir. İleride sistem canlıya geçtiğinde zorunlu tutabiliriz.
    return x_api_key

@app.get("/")
def home():
    return {"status": "online", "message": "Akilli Dönüstürücü Motoru Aktif"}


# ==========================================
# 1. MOTOR: PDF -> EXCEL (AKILLI TABLO ANALİZİ)
# ==========================================
@app.post("/convert/pdf-to-excel")
async def pdf_to_excel(
    file_url: str = Form(...),
    x_api_key: str = Header(None)
):
    verify_api_key(x_api_key)

    if file_url.startswith("//"):
        file_url = "https:" + file_url
        
    res = requests.get(file_url)
    if res.status_code != 200:
        raise HTTPException(status_code=400, detail="URL'den dosya indirilemedi.")
        
    pdf_bytes = res.content

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Veri Tablosu"

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            # Akıllı Analiz: Önce sıkı tablo yapısını (lattice) dene, bulamazsa serbest metin düzenini (stream) analiz et
            tables = page.extract_tables()
            
            if not tables:
                # Tablo çizgisi olmayan metin bazlı verileri yapılandırılmış olarak çıkar
                extracted_text = page.extract_text()
                if extracted_text:
                    for line in extracted_text.split("\n"):
                        ws.append([line])
            else:
                for table in tables:
                    for row_data in table:
                        # Boş ve hatalı hücreleri temizleyip ekle
                        cleaned_row = [cell.strip() if cell else "" for cell in row_data]
                        ws.append(cleaned_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        headers={"Content-Disposition": "attachment; filename=donusturulen_tablo.xlsx"},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================================
# 2. MOTOR: EXCEL -> PDF (SAYFAYA SIĞDIRMALI JİLET MİZANPAJ)
# ==========================================
@app.post("/convert/excel-to-pdf")
async def excel_to_pdf(
    file_url: str = Form(...),
    x_api_key: str = Header(None)
):
    verify_api_key(x_api_key)

    if file_url.startswith("//"):
        file_url = "https:" + file_url
        
    res = requests.get(file_url)
    if res.status_code != 200:
        raise HTTPException(status_code=400, detail="URL'den Excel dosyasi indirilemedi.")
        
    excel_bytes = res.content
    wb = openpyxl.load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    # Excel verisini oku
    data = []
    max_cols = 0
    for row in ws.iter_rows(values_only=True):
        row_data = [str(val) if val is not None else "" for val in row]
        if any(row_data): # Tamamen boş satırları atla
            data.append(row_data)
            if len(row_data) > max_cols:
                max_cols = len(row_data)

    if not data:
        raise HTTPException(status_code=400, detail="Excel dosyasi bos.")

    # PDF Mizanpaj Ayarları
    pdf_buffer = io.BytesIO()
    # Sütun sayısı çoksa yatay A4 (landscape), azsa dikey A4 kullan
    page_size = landscape(A4) if max_cols > 6 else A4
    
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=page_size,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]
    normal_style.fontSize = 8
    normal_style.leading = 10

    # Metinlerin taşmaması için hücre içlerini Paragraph objesine çevir
    table_data = []
    for row in data:
        formatted_row = []
        for cell in row:
            formatted_row.append(Paragraph(cell, normal_style))
        table_data.append(formatted_row)

    # Profesyonel Tablo Stili
    pdf_table = Table(table_data)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2C3E50")), # Koyu Lacivert Başlık
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")), # İnce Gri Çizgiler
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]), # Zebra Çizgisi
    ]))

    elements = [pdf_table]
    doc.build(elements)

    pdf_buffer.seek(0)

    return StreamingResponse(
        pdf_buffer,
        headers={"Content-Disposition": "attachment; filename=donusturulen_dokuman.pdf"},
        media_type="application/pdf"
    )
